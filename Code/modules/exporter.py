"""
exporter.py
===========
Exports KPI results to CSV and Excel (.xlsx).
"""

from __future__ import annotations
import csv
import os


def export_csv(results: list[dict], filepath: str) -> None:
    """Write KPI results list to a CSV file. O(n)."""
    if not results:
        return
    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=results[0].keys())
        writer.writeheader()
        writer.writerows(results)


def export_xlsx(results: list[dict], filepath: str) -> None:
    """Write KPI results list to an Excel file. O(n)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed — skipping xlsx export.")
        return

    if not results:
        return

    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI Results"

    headers = list(results[0].keys())
    header_fill = PatternFill("solid", fgColor="1A5E38")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    alt_fill = PatternFill("solid", fgColor="E3F0E8")
    for row_i, record in enumerate(results, 2):
        fill = alt_fill if row_i % 2 == 0 else PatternFill()
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_i, column=col, value=record[key])
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = max_len

    wb.save(filepath)
